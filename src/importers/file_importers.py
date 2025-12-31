"""
ANALYTICA File Importers
========================

Import data from various file formats:
- CSV (standard, Polish formats)
- Excel (xlsx, xls)
- JSON
- XML
- JPK (Polish tax files)
- MT940 (bank statements)
- QIF/OFX (financial exports)
- PDF (extraction)
"""

from typing import Any, Dict, List, Optional, Union, BinaryIO, TextIO
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from abc import ABC, abstractmethod
import csv
import json
import io
import re
import xml.etree.ElementTree as ET
from enum import Enum


# ============================================================
# BASE CLASSES
# ============================================================

class FileFormat(Enum):
    CSV = "csv"
    TSV = "tsv"
    EXCEL = "xlsx"
    EXCEL_OLD = "xls"
    JSON = "json"
    XML = "xml"
    JPK_VAT = "jpk_vat"
    JPK_FA = "jpk_fa"
    MT940 = "mt940"
    QIF = "qif"
    OFX = "ofx"
    PDF = "pdf"


@dataclass
class ImportResult:
    """Result of file import"""
    success: bool
    data: List[Dict]
    row_count: int
    columns: List[str]
    errors: List[Dict] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    metadata: Dict = field(default_factory=dict)
    source_format: FileFormat = FileFormat.CSV


class FileImporter(ABC):
    """Base class for file importers"""
    
    @abstractmethod
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import from file path"""
        pass
    
    @abstractmethod
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import from content string/bytes"""
        pass
    
    def _detect_encoding(self, content: bytes) -> str:
        """Detect file encoding"""
        # Check for BOM
        if content.startswith(b'\xef\xbb\xbf'):
            return 'utf-8-sig'
        if content.startswith(b'\xff\xfe'):
            return 'utf-16-le'
        if content.startswith(b'\xfe\xff'):
            return 'utf-16-be'
        
        # Try common encodings
        for encoding in ['utf-8', 'cp1250', 'iso-8859-2', 'cp1252']:
            try:
                content.decode(encoding)
                return encoding
            except UnicodeDecodeError:
                continue
        
        return 'utf-8'
    
    def _parse_decimal(self, value: str, polish_format: bool = True) -> Optional[Decimal]:
        """Parse decimal from string, handling Polish format"""
        if not value or not value.strip():
            return None
        
        value = value.strip()
        
        # Remove currency symbols and spaces
        value = re.sub(r'[PLNzłEUR€$\s]', '', value, flags=re.IGNORECASE)
        
        if polish_format:
            # Polish: 1 234,56 -> 1234.56
            value = value.replace(' ', '').replace(',', '.')
        
        try:
            return Decimal(value)
        except InvalidOperation:
            return None
    
    def _parse_date(self, value: str) -> Optional[date]:
        """Parse date from various formats"""
        if not value or not value.strip():
            return None
        
        value = value.strip()
        
        formats = [
            "%Y-%m-%d",      # ISO
            "%d.%m.%Y",      # Polish
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",      # US
            "%Y%m%d",        # Compact
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        
        return None


# ============================================================
# CSV IMPORTER
# ============================================================

class CSVImporter(FileImporter):
    """
    CSV/TSV file importer with automatic detection of:
    - Delimiter (comma, semicolon, tab)
    - Encoding
    - Decimal format (dot vs comma)
    - Date format
    """
    
    def __init__(
        self,
        delimiter: str = None,
        encoding: str = None,
        has_header: bool = True,
        skip_rows: int = 0,
        polish_format: bool = True
    ):
        self.delimiter = delimiter
        self.encoding = encoding
        self.has_header = has_header
        self.skip_rows = skip_rows
        self.polish_format = polish_format
    
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import CSV from file"""
        filepath = Path(filepath)
        
        with open(filepath, 'rb') as f:
            content = f.read()
        
        return self.import_content(content)
    
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import CSV from content"""
        # Handle bytes
        if isinstance(content, bytes):
            encoding = self.encoding or self._detect_encoding(content)
            content = content.decode(encoding)
        
        # Detect delimiter
        delimiter = self.delimiter or self._detect_delimiter(content)
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = list(reader)
        
        # Skip rows
        rows = rows[self.skip_rows:]
        
        if not rows:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": "Empty file"}]
            )
        
        # Extract headers
        if self.has_header:
            headers = [h.strip() for h in rows[0]]
            data_rows = rows[1:]
        else:
            headers = [f"col_{i}" for i in range(len(rows[0]))]
            data_rows = rows
        
        # Convert to dicts
        data = []
        errors = []
        
        for i, row in enumerate(data_rows):
            if len(row) != len(headers):
                errors.append({
                    "row": i + self.skip_rows + (2 if self.has_header else 1),
                    "error": f"Column count mismatch: expected {len(headers)}, got {len(row)}"
                })
                continue
            
            record = {}
            for j, (header, value) in enumerate(zip(headers, row)):
                record[header] = self._convert_value(value)
            
            data.append(record)
        
        return ImportResult(
            success=len(errors) == 0,
            data=data,
            row_count=len(data),
            columns=headers,
            errors=errors,
            source_format=FileFormat.CSV,
            metadata={
                "delimiter": delimiter,
                "encoding": self.encoding,
                "total_rows": len(rows)
            }
        )
    
    def _detect_delimiter(self, content: str) -> str:
        """Detect CSV delimiter"""
        first_lines = content[:2000]
        
        # Count occurrences
        counts = {
            ',': first_lines.count(','),
            ';': first_lines.count(';'),
            '\t': first_lines.count('\t'),
            '|': first_lines.count('|')
        }
        
        # Return most common
        return max(counts, key=counts.get)
    
    def _convert_value(self, value: str) -> Any:
        """Convert string value to appropriate type"""
        if not value or not value.strip():
            return None
        
        value = value.strip()
        
        # Try decimal
        decimal_val = self._parse_decimal(value, self.polish_format)
        if decimal_val is not None:
            return decimal_val
        
        # Try date
        date_val = self._parse_date(value)
        if date_val is not None:
            return date_val
        
        # Keep as string
        return value


# ============================================================
# EXCEL IMPORTER
# ============================================================

class ExcelImporter(FileImporter):
    """
    Excel file importer (xlsx, xls)
    
    Requires: openpyxl (xlsx), xlrd (xls)
    """
    
    def __init__(
        self,
        sheet_name: Union[str, int] = 0,
        has_header: bool = True,
        skip_rows: int = 0,
        polish_format: bool = True
    ):
        self.sheet_name = sheet_name
        self.has_header = has_header
        self.skip_rows = skip_rows
        self.polish_format = polish_format
    
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import Excel from file"""
        try:
            import openpyxl
        except ImportError:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": "openpyxl not installed"}]
            )
        
        filepath = Path(filepath)
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # Get sheet
            if isinstance(self.sheet_name, int):
                sheet = wb.worksheets[self.sheet_name]
            else:
                sheet = wb[self.sheet_name]
            
            # Read all rows
            rows = list(sheet.iter_rows(values_only=True))
            
            # Skip rows
            rows = rows[self.skip_rows:]
            
            if not rows:
                return ImportResult(
                    success=False,
                    data=[],
                    row_count=0,
                    columns=[],
                    errors=[{"error": "Empty sheet"}]
                )
            
            # Extract headers
            if self.has_header:
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                data_rows = rows[1:]
            else:
                headers = [f"col_{i}" for i in range(len(rows[0]))]
                data_rows = rows
            
            # Convert to dicts
            data = []
            for row in data_rows:
                if all(v is None for v in row):
                    continue  # Skip empty rows
                
                record = {}
                for header, value in zip(headers, row):
                    record[header] = self._convert_excel_value(value)
                
                data.append(record)
            
            return ImportResult(
                success=True,
                data=data,
                row_count=len(data),
                columns=headers,
                source_format=FileFormat.EXCEL,
                metadata={
                    "sheet_name": sheet.title,
                    "sheets": wb.sheetnames
                }
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": str(e)}]
            )
    
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import Excel from bytes"""
        if isinstance(content, str):
            content = content.encode()
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            
            # Get sheet
            if isinstance(self.sheet_name, int):
                sheet = wb.worksheets[self.sheet_name]
            else:
                sheet = wb[self.sheet_name]
            
            rows = list(sheet.iter_rows(values_only=True))
            rows = rows[self.skip_rows:]
            
            if self.has_header:
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                data_rows = rows[1:]
            else:
                headers = [f"col_{i}" for i in range(len(rows[0]))]
                data_rows = rows
            
            data = []
            for row in data_rows:
                if all(v is None for v in row):
                    continue
                
                record = {}
                for header, value in zip(headers, row):
                    record[header] = self._convert_excel_value(value)
                
                data.append(record)
            
            return ImportResult(
                success=True,
                data=data,
                row_count=len(data),
                columns=headers,
                source_format=FileFormat.EXCEL
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": str(e)}]
            )
    
    def _convert_excel_value(self, value: Any) -> Any:
        """Convert Excel cell value"""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        return str(value).strip()


# ============================================================
# JSON IMPORTER
# ============================================================

class JSONImporter(FileImporter):
    """JSON file importer"""
    
    def __init__(self, data_path: str = None):
        """
        Args:
            data_path: JSON path to data array (e.g., "results.data")
        """
        self.data_path = data_path
    
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import JSON from file"""
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        return self.import_content(content)
    
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import JSON from content"""
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        
        try:
            data = json.loads(content)
            
            # Navigate to data path
            if self.data_path:
                for key in self.data_path.split('.'):
                    if isinstance(data, dict):
                        data = data.get(key, [])
                    elif isinstance(data, list) and key.isdigit():
                        data = data[int(key)]
                    else:
                        break
            
            # Ensure data is a list
            if isinstance(data, dict):
                data = [data]
            elif not isinstance(data, list):
                data = []
            
            # Extract columns
            columns = []
            if data:
                columns = list(data[0].keys()) if isinstance(data[0], dict) else []
            
            return ImportResult(
                success=True,
                data=data,
                row_count=len(data),
                columns=columns,
                source_format=FileFormat.JSON
            )
            
        except json.JSONDecodeError as e:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": f"JSON parse error: {e}"}]
            )


# ============================================================
# XML IMPORTER
# ============================================================

class XMLImporter(FileImporter):
    """XML file importer"""
    
    def __init__(
        self,
        record_tag: str = None,
        namespaces: Dict[str, str] = None
    ):
        """
        Args:
            record_tag: XML tag for each record (e.g., "item", "row")
            namespaces: XML namespaces dict
        """
        self.record_tag = record_tag
        self.namespaces = namespaces or {}
    
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import XML from file"""
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        return self.import_content(content)
    
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import XML from content"""
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        
        try:
            root = ET.fromstring(content)
            
            # Find records
            if self.record_tag:
                records = root.findall(f".//{self.record_tag}", self.namespaces)
            else:
                # Use first level children
                records = list(root)
            
            data = []
            columns = set()
            
            for record in records:
                item = self._element_to_dict(record)
                data.append(item)
                columns.update(item.keys())
            
            return ImportResult(
                success=True,
                data=data,
                row_count=len(data),
                columns=list(columns),
                source_format=FileFormat.XML,
                metadata={
                    "root_tag": root.tag,
                    "record_tag": self.record_tag
                }
            )
            
        except ET.ParseError as e:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": f"XML parse error: {e}"}]
            )
    
    def _element_to_dict(self, element: ET.Element) -> Dict:
        """Convert XML element to dict"""
        result = {}
        
        # Add attributes
        for key, value in element.attrib.items():
            result[f"@{key}"] = value
        
        # Add children
        for child in element:
            tag = child.tag.split('}')[-1]  # Remove namespace
            
            if len(child) > 0:
                # Has children - recurse
                result[tag] = self._element_to_dict(child)
            else:
                # Leaf node
                result[tag] = child.text.strip() if child.text else None
        
        # If no children, use text content
        if not result and element.text:
            return element.text.strip()
        
        return result


# ============================================================
# JPK IMPORTER (Polish Tax Files)
# ============================================================

class JPKImporter(FileImporter):
    """
    JPK (Jednolity Plik Kontrolny) Importer
    
    Supports:
    - JPK_VAT (VAT records)
    - JPK_FA (Invoices)
    - JPK_KR (Accounting books)
    """
    
    JPK_NAMESPACES = {
        'jpk': 'http://crd.gov.pl/wzor/2022/02/17/11148/',
        'etd': 'http://crd.gov.pl/xml/schematy/dziedzinowe/mf/2016/01/25/eD/DefinicjeTypy/'
    }
    
    def import_file(self, filepath: Union[str, Path]) -> ImportResult:
        """Import JPK from file"""
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        return self.import_content(content)
    
    def import_content(self, content: Union[str, bytes]) -> ImportResult:
        """Import JPK from content"""
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        
        try:
            root = ET.fromstring(content)
            
            # Detect JPK type
            jpk_type = self._detect_jpk_type(root)
            
            if jpk_type == "JPK_VAT":
                return self._parse_jpk_vat(root)
            elif jpk_type == "JPK_FA":
                return self._parse_jpk_fa(root)
            else:
                return ImportResult(
                    success=False,
                    data=[],
                    row_count=0,
                    columns=[],
                    errors=[{"error": f"Unsupported JPK type: {jpk_type}"}]
                )
            
        except ET.ParseError as e:
            return ImportResult(
                success=False,
                data=[],
                row_count=0,
                columns=[],
                errors=[{"error": f"XML parse error: {e}"}]
            )
    
    def _detect_jpk_type(self, root: ET.Element) -> str:
        """Detect JPK file type"""
        # Try to find KodFormularza
        for elem in root.iter():
            if 'KodFormularza' in elem.tag:
                return elem.text or elem.get('kodSystemowy', '')
        
        return "UNKNOWN"
    
    def _parse_jpk_vat(self, root: ET.Element) -> ImportResult:
        """Parse JPK_VAT file"""
        records = []
        
        # Find all SprzedazWiersz (sales) and ZakupWiersz (purchases)
        for tag in ['SprzedazWiersz', 'ZakupWiersz']:
            for elem in root.iter():
                if tag in elem.tag:
                    record = {
                        'type': 'sprzedaz' if 'Sprzedaz' in tag else 'zakup'
                    }
                    
                    for child in elem:
                        tag_name = child.tag.split('}')[-1]
                        record[tag_name] = child.text
                    
                    records.append(record)
        
        columns = list(records[0].keys()) if records else []
        
        return ImportResult(
            success=True,
            data=records,
            row_count=len(records),
            columns=columns,
            source_format=FileFormat.JPK_VAT,
            metadata={"jpk_type": "JPK_VAT"}
        )
    
    def _parse_jpk_fa(self, root: ET.Element) -> ImportResult:
        """Parse JPK_FA (invoices) file"""
        records = []
        
        for elem in root.iter():
            if 'Faktura' in elem.tag:
                record = {}
                
                for child in elem:
                    tag_name = child.tag.split('}')[-1]
                    record[tag_name] = child.text
                
                records.append(record)
        
        columns = list(records[0].keys()) if records else []
        
        return ImportResult(
            success=True,
            data=records,
            row_count=len(records),
            columns=columns,
            source_format=FileFormat.JPK_FA,
            metadata={"jpk_type": "JPK_FA"}
        )


# ============================================================
# FACTORY
# ============================================================

def create_importer(format: Union[str, FileFormat], **kwargs) -> FileImporter:
    """
    Factory function to create file importer
    
    Args:
        format: File format (csv, xlsx, json, xml, jpk, mt940)
        **kwargs: Format-specific options
    
    Returns:
        Configured importer instance
    """
    if isinstance(format, str):
        format = format.lower()
    else:
        format = format.value
    
    importers = {
        'csv': CSVImporter,
        'tsv': lambda **kw: CSVImporter(delimiter='\t', **kw),
        'xlsx': ExcelImporter,
        'xls': ExcelImporter,
        'excel': ExcelImporter,
        'json': JSONImporter,
        'xml': XMLImporter,
        'jpk': JPKImporter,
        'jpk_vat': JPKImporter,
        'jpk_fa': JPKImporter,
    }
    
    if format not in importers:
        raise ValueError(f"Unsupported format: {format}. Supported: {list(importers.keys())}")
    
    return importers[format](**kwargs)


def import_file(
    filepath: Union[str, Path],
    format: str = None,
    **kwargs
) -> ImportResult:
    """
    Quick function to import file
    
    Auto-detects format from extension if not specified
    """
    filepath = Path(filepath)
    
    if not format:
        ext = filepath.suffix.lower().lstrip('.')
        format = ext
    
    importer = create_importer(format, **kwargs)
    return importer.import_file(filepath)


__all__ = [
    'FileFormat',
    'ImportResult',
    'FileImporter',
    'CSVImporter',
    'ExcelImporter',
    'JSONImporter',
    'XMLImporter',
    'JPKImporter',
    'create_importer',
    'import_file'
]
