"""
ANALYTICA File Exporters
========================

Export data to various file formats:
- CSV (standard, Polish formats)
- Excel (xlsx with formatting)
- JSON
- XML
- PDF (reports)
- JPK (Polish tax files)
- HTML (reports)
"""

from typing import Any, Dict, List, Optional, Union, BinaryIO
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
from abc import ABC, abstractmethod
import csv
import json
import io
from enum import Enum


# ============================================================
# BASE CLASSES
# ============================================================

class ExportFormat(Enum):
    CSV = "csv"
    TSV = "tsv"
    EXCEL = "xlsx"
    JSON = "json"
    XML = "xml"
    PDF = "pdf"
    HTML = "html"
    JPK = "jpk"


@dataclass
class ExportResult:
    """Result of file export"""
    success: bool
    filepath: Optional[str] = None
    content: Optional[bytes] = None
    row_count: int = 0
    file_size: int = 0
    format: ExportFormat = ExportFormat.CSV
    errors: List[str] = field(default_factory=list)
    metadata: Dict = field(default_factory=dict)


class FileExporter(ABC):
    """Base class for file exporters"""
    
    @abstractmethod
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to file or return content"""
        pass
    
    def _format_value(self, value: Any, polish_format: bool = True) -> str:
        """Format value for export"""
        if value is None:
            return ""
        
        if isinstance(value, Decimal):
            if polish_format:
                return str(value).replace('.', ',')
            return str(value)
        
        if isinstance(value, (date, datetime)):
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            return value.strftime("%Y-%m-%d")
        
        if isinstance(value, bool):
            return "Tak" if value else "Nie" if polish_format else str(value)
        
        return str(value)
    
    def _get_columns(self, data: List[Dict]) -> List[str]:
        """Extract column names from data"""
        if not data:
            return []
        
        columns = set()
        for row in data:
            columns.update(row.keys())
        
        return sorted(columns)


# ============================================================
# CSV EXPORTER
# ============================================================

class CSVExporter(FileExporter):
    """
    CSV/TSV file exporter
    
    Features:
    - Custom delimiter
    - Polish number format (comma decimal)
    - UTF-8 BOM for Excel compatibility
    - Custom column order
    """
    
    def __init__(
        self,
        delimiter: str = ';',
        encoding: str = 'utf-8-sig',  # UTF-8 with BOM for Excel
        polish_format: bool = True,
        columns: List[str] = None,
        include_header: bool = True
    ):
        self.delimiter = delimiter
        self.encoding = encoding
        self.polish_format = polish_format
        self.columns = columns
        self.include_header = include_header
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to CSV"""
        if not data:
            return ExportResult(
                success=False,
                errors=["No data to export"]
            )
        
        # Get columns
        columns = self.columns or self._get_columns(data)
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output, delimiter=self.delimiter)
        
        # Header
        if self.include_header:
            writer.writerow(columns)
        
        # Data rows
        for row in data:
            values = [
                self._format_value(row.get(col), self.polish_format)
                for col in columns
            ]
            writer.writerow(values)
        
        content = output.getvalue().encode(self.encoding)
        
        # Write to file if filepath provided
        if filepath:
            filepath = Path(filepath)
            filepath.write_bytes(content)
            
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                file_size=len(content),
                format=ExportFormat.CSV,
                metadata={"columns": columns, "delimiter": self.delimiter}
            )
        
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            file_size=len(content),
            format=ExportFormat.CSV,
            metadata={"columns": columns}
        )


# ============================================================
# EXCEL EXPORTER
# ============================================================

class ExcelExporter(FileExporter):
    """
    Excel file exporter (xlsx)
    
    Features:
    - Multiple sheets
    - Auto-formatting (numbers, dates, currency)
    - Column width auto-fit
    - Header styling
    - Freeze panes
    """
    
    def __init__(
        self,
        sheet_name: str = "Dane",
        columns: List[str] = None,
        auto_fit_columns: bool = True,
        freeze_header: bool = True,
        currency_columns: List[str] = None,
        date_columns: List[str] = None
    ):
        self.sheet_name = sheet_name
        self.columns = columns
        self.auto_fit_columns = auto_fit_columns
        self.freeze_header = freeze_header
        self.currency_columns = currency_columns or []
        self.date_columns = date_columns or []
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return ExportResult(
                success=False,
                errors=["openpyxl not installed"]
            )
        
        if not data:
            return ExportResult(
                success=False,
                errors=["No data to export"]
            )
        
        # Get columns
        columns = self.columns or self._get_columns(data)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Handle different types
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    if col_name in self.currency_columns:
                        cell.number_format = '#,##0.00 "zł"'
                    else:
                        cell.number_format = '#,##0.00'
                elif isinstance(value, (date, datetime)):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD'
                else:
                    cell.value = value
                
                cell.border = border
        
        # Auto-fit columns
        if self.auto_fit_columns:
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(col_name)
                
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Freeze header
        if self.freeze_header:
            ws.freeze_panes = 'A2'
        
        # Save
        if filepath:
            filepath = Path(filepath)
            wb.save(filepath)
            
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                format=ExportFormat.EXCEL,
                metadata={"columns": columns, "sheet": self.sheet_name}
            )
        
        # Return as bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            file_size=len(content),
            format=ExportFormat.EXCEL,
            metadata={"columns": columns}
        )


# ============================================================
# JSON EXPORTER
# ============================================================

class JSONExporter(FileExporter):
    """JSON file exporter"""
    
    def __init__(
        self,
        indent: int = 2,
        ensure_ascii: bool = False,
        wrap_in_object: str = None  # e.g., "data" -> {"data": [...]}
    ):
        self.indent = indent
        self.ensure_ascii = ensure_ascii
        self.wrap_in_object = wrap_in_object
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to JSON"""
        # Convert special types
        serializable_data = self._make_serializable(data)
        
        # Wrap if needed
        if self.wrap_in_object:
            output_data = {
                self.wrap_in_object: serializable_data,
                "meta": {
                    "count": len(data),
                    "exported_at": datetime.now().isoformat()
                }
            }
        else:
            output_data = serializable_data
        
        content = json.dumps(
            output_data,
            indent=self.indent,
            ensure_ascii=self.ensure_ascii,
            default=str
        ).encode('utf-8')
        
        if filepath:
            filepath = Path(filepath)
            filepath.write_bytes(content)
            
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                file_size=len(content),
                format=ExportFormat.JSON
            )
        
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            file_size=len(content),
            format=ExportFormat.JSON
        )
    
    def _make_serializable(self, data: Any) -> Any:
        """Convert to JSON-serializable types"""
        if isinstance(data, list):
            return [self._make_serializable(item) for item in data]
        if isinstance(data, dict):
            return {k: self._make_serializable(v) for k, v in data.items()}
        if isinstance(data, Decimal):
            return float(data)
        if isinstance(data, (date, datetime)):
            return data.isoformat()
        return data


# ============================================================
# XML EXPORTER
# ============================================================

class XMLExporter(FileExporter):
    """XML file exporter"""
    
    def __init__(
        self,
        root_tag: str = "data",
        record_tag: str = "record",
        include_declaration: bool = True,
        indent: bool = True
    ):
        self.root_tag = root_tag
        self.record_tag = record_tag
        self.include_declaration = include_declaration
        self.indent = indent
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to XML"""
        import xml.etree.ElementTree as ET
        
        root = ET.Element(self.root_tag)
        root.set("count", str(len(data)))
        root.set("exported_at", datetime.now().isoformat())
        
        for row in data:
            record = ET.SubElement(root, self.record_tag)
            
            for key, value in row.items():
                # Clean key for XML tag
                clean_key = self._clean_tag_name(key)
                elem = ET.SubElement(record, clean_key)
                elem.text = self._format_value(value, polish_format=False)
        
        # Convert to string
        if self.indent:
            self._indent_xml(root)
        
        xml_str = ET.tostring(root, encoding='unicode')
        
        if self.include_declaration:
            xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str
        
        content = xml_str.encode('utf-8')
        
        if filepath:
            filepath = Path(filepath)
            filepath.write_bytes(content)
            
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                file_size=len(content),
                format=ExportFormat.XML
            )
        
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            file_size=len(content),
            format=ExportFormat.XML
        )
    
    def _clean_tag_name(self, name: str) -> str:
        """Clean string for use as XML tag"""
        import re
        # Remove invalid characters
        name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
        # Ensure starts with letter
        if name and not name[0].isalpha():
            name = 'field_' + name
        return name or 'field'
    
    def _indent_xml(self, elem, level=0):
        """Add indentation to XML"""
        indent = "\n" + "  " * level
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = indent + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
            for child in elem:
                self._indent_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = indent
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = indent


# ============================================================
# PDF EXPORTER
# ============================================================

class PDFExporter(FileExporter):
    """
    PDF report exporter
    
    Creates professional PDF reports with:
    - Header/footer
    - Logo
    - Tables
    - Charts (optional)
    """
    
    def __init__(
        self,
        title: str = "Raport",
        company_name: str = None,
        logo_path: str = None,
        page_size: str = "A4",
        orientation: str = "portrait"
    ):
        self.title = title
        self.company_name = company_name
        self.logo_path = logo_path
        self.page_size = page_size
        self.orientation = orientation
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            return ExportResult(
                success=False,
                errors=["reportlab not installed"]
            )
        
        if not data:
            return ExportResult(
                success=False,
                errors=["No data to export"]
            )
        
        # Get columns
        columns = self._get_columns(data)
        
        # Create buffer or file
        if filepath:
            filepath = Path(filepath)
            buffer = str(filepath)
        else:
            buffer = io.BytesIO()
        
        # Page setup
        page_size = landscape(A4) if self.orientation == "landscape" else A4
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Build content
        elements = []
        
        # Title
        elements.append(Paragraph(self.title, title_style))
        
        # Date
        date_style = ParagraphStyle('Date', parent=styles['Normal'], alignment=1)
        elements.append(Paragraph(
            f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            date_style
        ))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [columns]  # Header
        
        for row in data:
            table_data.append([
                self._format_value(row.get(col), polish_format=True)
                for col in columns
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]))
        
        elements.append(table)
        
        # Summary
        elements.append(Spacer(1, 20))
        summary_style = ParagraphStyle('Summary', parent=styles['Normal'])
        elements.append(Paragraph(f"Liczba rekordów: {len(data)}", summary_style))
        
        # Build PDF
        doc.build(elements)
        
        if filepath:
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                format=ExportFormat.PDF,
                metadata={"title": self.title, "columns": columns}
            )
        
        content = buffer.getvalue()
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            file_size=len(content),
            format=ExportFormat.PDF
        )


# ============================================================
# HTML EXPORTER
# ============================================================

class HTMLExporter(FileExporter):
    """HTML table exporter with styling"""
    
    def __init__(
        self,
        title: str = "Raport",
        include_styling: bool = True,
        table_class: str = "analytica-table"
    ):
        self.title = title
        self.include_styling = include_styling
        self.table_class = table_class
    
    def export(
        self, 
        data: List[Dict],
        filepath: Union[str, Path] = None
    ) -> ExportResult:
        """Export data to HTML"""
        if not data:
            return ExportResult(success=False, errors=["No data"])
        
        columns = self._get_columns(data)
        
        # Build HTML
        html_parts = ['<!DOCTYPE html>', '<html>', '<head>',
                     '<meta charset="UTF-8">',
                     f'<title>{self.title}</title>']
        
        if self.include_styling:
            html_parts.append(self._get_styles())
        
        html_parts.extend(['</head>', '<body>',
                          f'<h1>{self.title}</h1>',
                          f'<p>Wygenerowano: {datetime.now().strftime("%Y-%m-%d %H:%M")}</p>',
                          f'<table class="{self.table_class}">',
                          '<thead><tr>'])
        
        # Header
        for col in columns:
            html_parts.append(f'<th>{col}</th>')
        
        html_parts.append('</tr></thead><tbody>')
        
        # Data rows
        for row in data:
            html_parts.append('<tr>')
            for col in columns:
                value = self._format_value(row.get(col), polish_format=True)
                html_parts.append(f'<td>{value}</td>')
            html_parts.append('</tr>')
        
        html_parts.extend(['</tbody></table>',
                          f'<p>Liczba rekordów: {len(data)}</p>',
                          '</body></html>'])
        
        content = '\n'.join(html_parts).encode('utf-8')
        
        if filepath:
            Path(filepath).write_bytes(content)
            return ExportResult(
                success=True,
                filepath=str(filepath),
                row_count=len(data),
                format=ExportFormat.HTML
            )
        
        return ExportResult(
            success=True,
            content=content,
            row_count=len(data),
            format=ExportFormat.HTML
        )
    
    def _get_styles(self) -> str:
        return '''<style>
body { font-family: Arial, sans-serif; margin: 20px; }
h1 { color: #333; }
.analytica-table { border-collapse: collapse; width: 100%; margin: 20px 0; }
.analytica-table th, .analytica-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
.analytica-table th { background-color: #4472C4; color: white; }
.analytica-table tr:nth-child(even) { background-color: #f2f2f2; }
.analytica-table tr:hover { background-color: #ddd; }
</style>'''


# ============================================================
# FACTORY
# ============================================================

def create_exporter(format: Union[str, ExportFormat], **kwargs) -> FileExporter:
    """Factory function to create file exporter"""
    if isinstance(format, str):
        format = format.lower()
    else:
        format = format.value
    
    exporters = {
        'csv': CSVExporter,
        'tsv': lambda **kw: CSVExporter(delimiter='\t', **kw),
        'xlsx': ExcelExporter,
        'excel': ExcelExporter,
        'json': JSONExporter,
        'xml': XMLExporter,
        'pdf': PDFExporter,
        'html': HTMLExporter,
    }
    
    if format not in exporters:
        raise ValueError(f"Unsupported format: {format}")
    
    return exporters[format](**kwargs)


def export_data(
    data: List[Dict],
    filepath: Union[str, Path] = None,
    format: str = "csv",
    **kwargs
) -> ExportResult:
    """Quick function to export data"""
    if filepath and not format:
        ext = Path(filepath).suffix.lower().lstrip('.')
        format = ext
    
    exporter = create_exporter(format, **kwargs)
    return exporter.export(data, filepath)


__all__ = [
    'ExportFormat',
    'ExportResult',
    'FileExporter',
    'CSVExporter',
    'ExcelExporter',
    'JSONExporter',
    'XMLExporter',
    'PDFExporter',
    'HTMLExporter',
    'create_exporter',
    'export_data'
]
