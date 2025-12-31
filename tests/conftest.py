"""
ANALYTICA Tests - Configuration and Fixtures
=============================================
"""

import pytest
import asyncio
import json
import os
from pathlib import Path
from typing import Dict, Any, Generator
from datetime import datetime, date
from decimal import Decimal
import tempfile

# Add src to path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from dsl.core.parser import Pipeline, PipelineBuilder, PipelineContext, parse, execute
from dsl.atoms.implementations import *  # Register atoms


# ============================================================
# PYTEST CONFIGURATION
# ============================================================

def pytest_configure(config):
    """Configure pytest"""
    config.addinivalue_line("markers", "e2e: End-to-end tests")
    config.addinivalue_line("markers", "unit: Unit tests")
    config.addinivalue_line("markers", "integration: Integration tests")
    config.addinivalue_line("markers", "slow: Slow tests")


@pytest.fixture(scope="session")
def event_loop():
    """Create event loop for async tests"""
    loop = asyncio.get_event_loop_policy().new_event_loop()
    yield loop
    loop.close()


# ============================================================
# SAMPLE DATA FIXTURES
# ============================================================

@pytest.fixture
def sample_sales_data() -> list:
    """Sample sales data"""
    return [
        {"id": 1, "date": "2024-01-15", "amount": 1500.00, "category": "Product A", "region": "PL"},
        {"id": 2, "date": "2024-01-20", "amount": 2300.00, "category": "Product B", "region": "PL"},
        {"id": 3, "date": "2024-02-10", "amount": 1800.00, "category": "Product A", "region": "DE"},
        {"id": 4, "date": "2024-02-15", "amount": 3200.00, "category": "Product C", "region": "PL"},
        {"id": 5, "date": "2024-03-01", "amount": 2100.00, "category": "Product B", "region": "DE"},
        {"id": 6, "date": "2024-03-10", "amount": 4500.00, "category": "Product A", "region": "PL"},
    ]


@pytest.fixture
def sample_budget_data() -> dict:
    """Sample budget data"""
    return {
        "budget_id": "budget_2024_q1",
        "name": "Budżet Q1 2024",
        "period_start": "2024-01-01",
        "period_end": "2024-03-31",
        "categories": [
            {"name": "Wynagrodzenia", "planned": 150000, "actual": 148500},
            {"name": "Marketing", "planned": 30000, "actual": 32500},
            {"name": "IT", "planned": 25000, "actual": 23000},
            {"name": "Biuro", "planned": 15000, "actual": 14200},
            {"name": "Podróże", "planned": 10000, "actual": 8500},
        ],
        "total_planned": 230000,
        "total_actual": 226700
    }


@pytest.fixture
def sample_investment_data() -> dict:
    """Sample investment data"""
    return {
        "name": "Nowa linia produkcyjna",
        "initial_investment": 500000,
        "discount_rate": 0.12,
        "investment_type": "capex",
        "cash_flows": [
            {"period": 1, "amount": 150000, "description": "Rok 1"},
            {"period": 2, "amount": 180000, "description": "Rok 2"},
            {"period": 3, "amount": 200000, "description": "Rok 3"},
            {"period": 4, "amount": 220000, "description": "Rok 4"},
            {"period": 5, "amount": 250000, "description": "Rok 5"},
        ]
    }


@pytest.fixture
def sample_forecast_data() -> list:
    """Sample time series data for forecasting"""
    return [
        {"date": "2024-01-01", "value": 100},
        {"date": "2024-01-02", "value": 105},
        {"date": "2024-01-03", "value": 102},
        {"date": "2024-01-04", "value": 110},
        {"date": "2024-01-05", "value": 108},
        {"date": "2024-01-06", "value": 115},
        {"date": "2024-01-07", "value": 112},
    ]


@pytest.fixture
def sample_transactions() -> list:
    """Sample financial transactions"""
    return [
        {"id": "TRX001", "date": "2024-01-15", "amount": 1500.00, "type": "income", "category": "Sprzedaż"},
        {"id": "TRX002", "date": "2024-01-16", "amount": -500.00, "type": "expense", "category": "Marketing"},
        {"id": "TRX003", "date": "2024-01-17", "amount": -1200.00, "type": "expense", "category": "Wynagrodzenia"},
        {"id": "TRX004", "date": "2024-01-18", "amount": 2300.00, "type": "income", "category": "Sprzedaż"},
        {"id": "TRX005", "date": "2024-01-19", "amount": -300.00, "type": "expense", "category": "IT"},
    ]


# ============================================================
# FILE FIXTURES
# ============================================================

@pytest.fixture
def temp_csv_file(sample_sales_data) -> Generator[str, None, None]:
    """Create temporary CSV file with sample data"""
    import csv
    
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        writer = csv.DictWriter(f, fieldnames=sample_sales_data[0].keys())
        writer.writeheader()
        writer.writerows(sample_sales_data)
        temp_path = f.name
    
    yield temp_path
    
    os.unlink(temp_path)


@pytest.fixture
def temp_json_file(sample_sales_data) -> Generator[str, None, None]:
    """Create temporary JSON file with sample data"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(sample_sales_data, f)
        temp_path = f.name
    
    yield temp_path
    
    os.unlink(temp_path)


@pytest.fixture
def temp_excel_file(sample_sales_data) -> Generator[str, None, None]:
    """Create temporary Excel file with sample data"""
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        headers = list(sample_sales_data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, item in enumerate(sample_sales_data, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item[key])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        wb.save(temp_path)
        
        yield temp_path
        
        os.unlink(temp_path)
    except ImportError:
        pytest.skip("openpyxl not installed")


# ============================================================
# MT940 BANKING FIXTURES
# ============================================================

@pytest.fixture
def sample_mt940_content() -> str:
    """Sample MT940 bank statement"""
    return """{1:F01INGBPLPWXXX0000000000}{2:O9401200240115INGBPLPWXXXX00000000002401151200N}{4:
:20:STMT240115
:25:PL12345678901234567890123456
:28C:1/1
:60F:C240114PLN10000,00
:61:2401150115C1500,00NTRFNONREF//TRX001
Wpłata od klienta ABC
:61:2401160116D500,00NTRFNONREF//TRX002
Opłata za usługi marketingowe
:61:2401170117D1200,00NTRFNONREF//TRX003
Wypłata wynagrodzeń
:62F:C240117PLN9800,00
-}"""


@pytest.fixture
def sample_jpk_vat() -> str:
    """Sample JPK_VAT XML structure"""
    return """<?xml version="1.0" encoding="UTF-8"?>
<JPK xmlns="http://crd.gov.pl/wzor/2022/02/17/11148/">
    <Naglowek>
        <KodFormularza kodSystemowy="JPK_VAT (3)" wersjaSchemy="1-0">JPK_VAT</KodFormularza>
        <WariantFormularza>3</WariantFormularza>
        <CelZlozenia>1</CelZlozenia>
        <DataWytworzeniaJPK>2024-01-20T10:00:00</DataWytworzeniaJPK>
        <DataOd>2024-01-01</DataOd>
        <DataDo>2024-01-31</DataDo>
    </Naglowek>
    <Podmiot1>
        <NIP>1234567890</NIP>
        <PelnaNazwa>Test Company Sp. z o.o.</PelnaNazwa>
    </Podmiot1>
</JPK>"""


# ============================================================
# PIPELINE FIXTURES
# ============================================================

@pytest.fixture
def pipeline_context() -> PipelineContext:
    """Create fresh pipeline context"""
    return PipelineContext(
        variables={},
        domain="planbudzetu.pl"
    )


@pytest.fixture
def pipeline_context_with_data(sample_sales_data) -> PipelineContext:
    """Create pipeline context with sample data"""
    ctx = PipelineContext(
        variables={"year": 2024},
        domain="planbudzetu.pl"
    )
    ctx.set_data(sample_sales_data)
    return ctx


@pytest.fixture
def pipeline_builder() -> PipelineBuilder:
    """Create fresh pipeline builder"""
    return PipelineBuilder(domain="planbudzetu.pl")


# ============================================================
# API CLIENT FIXTURES
# ============================================================

@pytest.fixture
def mock_api_response():
    """Mock API response factory"""
    def _create_response(data: dict, status_code: int = 200):
        class MockResponse:
            def __init__(self, data, status_code):
                self._data = data
                self.status_code = status_code
            
            def json(self):
                return self._data
            
            def raise_for_status(self):
                if self.status_code >= 400:
                    raise Exception(f"HTTP {self.status_code}")
        
        return MockResponse(data, status_code)
    
    return _create_response


@pytest.fixture
def ifirma_credentials() -> dict:
    """Mock iFirma API credentials"""
    return {
        "api_key": "test_api_key_12345",
        "username": "test@example.com",
        "key_name": "faktura"
    }


@pytest.fixture
def fakturownia_credentials() -> dict:
    """Mock Fakturownia API credentials"""
    return {
        "api_token": "test_token_xyz",
        "account_prefix": "testaccount"
    }


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def create_test_pipeline(dsl: str) -> Any:
    """Helper to create and parse a test pipeline"""
    return parse(dsl)


def execute_test_pipeline(dsl: str, input_data: Any = None, variables: dict = None) -> PipelineContext:
    """Helper to execute a test pipeline"""
    ctx = PipelineContext(variables=variables or {})
    if input_data:
        ctx.set_data(input_data)
    return execute(dsl, ctx)


# Export for use in tests
__all__ = [
    'sample_sales_data',
    'sample_budget_data', 
    'sample_investment_data',
    'sample_forecast_data',
    'sample_transactions',
    'temp_csv_file',
    'temp_json_file',
    'sample_mt940_content',
    'sample_jpk_vat',
    'pipeline_context',
    'pipeline_builder',
    'create_test_pipeline',
    'execute_test_pipeline'
]
